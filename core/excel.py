import os
import logging
from dataclasses import dataclass
from openpyxl import load_workbook
from django.db import transaction
from core.models import (
    Class,
    Subclass,
    Treatment,
    Reference,
    Compound,
    FormulaMass,
    ExcelUpload,
)
from core.utils import generate_and_save_molecule_image, clear_data, add_user_event


logger = logging.getLogger(__name__)


@dataclass
class FormulaMz:
    formula: str
    mz: str


@dataclass
class Columns:
    name: str
    origin: str
    clas: str
    subclass: str
    treatment: str
    type: str
    mode: str
    neutral: str
    mz: str
    reference: str
    smile: str
    notes: str
    formula_mz: list

    def valid(self):
        if None in [
            self.name,
            self.clas,
            self.subclass,
            self.type,
            self.mode,
            self.neutral,
            self.mz,
            self.smile,
            self.reference,
            self.notes,
        ]:
            return False
        return True


def get_columns(ws):
    for r in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        row = r

    col_nums = {}
    for i, v in enumerate(row):
        if v is None:
            continue
        col_nums[str(v).strip().lower()] = i

    formula_mz = []
    for i in range(1, 11):
        formula_name = f"fragment {i}"
        mz_name = f"m/z fragment {i}"
        formula = col_nums.get(formula_name)
        mz = col_nums.get(mz_name)
        formula_mz.append(FormulaMz(formula=formula, mz=mz))

    columns = Columns(
        clas=col_nums.get("compound class"),
        subclass=col_nums.get("subclass"),
        treatment=col_nums.get("treatment"),
        origin=col_nums.get("parent compound"),
        type=col_nums.get("type"),
        mode=col_nums.get("ionization mode"),
        name=col_nums.get("compound"),
        neutral=col_nums.get("molecular formula [m]"),
        mz=col_nums.get("m/z ion"),
        reference=col_nums.get("references"),
        smile=col_nums.get("smile neutral formula"),
        notes=col_nums.get("notes"),
        formula_mz=formula_mz,
    )

    if not columns.valid():
        raise Exception(f"Missing required columns in Excel file: {columns} {col_nums}")

    return columns


def field(row, col):
    return str(row[col]).strip() if row[col] else ""


def split_list(text):
    return [name.strip() for name in text.split(";") if name.strip()]


def create_formula_mass_models(row, columns: Columns):
    """Create FormulaMass models for F1-F10 columns and return list of created objects"""
    formula_mass_objects = []

    for formula_mz in columns.formula_mz:
        formula = formula_mz.formula
        mz = formula_mz.mz
        if formula is None or mz is None:
            continue

        formula_value = field(row, formula)
        mz_value = str(row[mz]).strip()  # field(row, mz)
        if not (formula_value and mz_value):
            continue

        formula_mass_obj, created = FormulaMass.objects.get_or_create(
            formula=formula_value, mass=mz_value
        )
        formula_mass_objects.append(formula_mass_obj)
        if created:
            logger.info(
                f"Created FormulaMass: {formula_mass_obj.formula} - {formula_mass_obj.mass}"
            )

    return formula_mass_objects


def import_excel_data(ws, skip_images=False):
    columns = get_columns(ws)

    classes = {}
    subclasses = {}
    treatments = {"-": None, "": None}
    references = {"-": None, "": None}
    originals = {}
    tps = {}
    total_count = 0

    # Process all rows
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue

        c = field(r, columns.clas)
        if c and c not in classes:
            obj, created = Class.objects.get_or_create(name=c)
            classes[c] = obj
            if created:
                logger.info(f"Created Class: {obj}")
                total_count += 1

        s = field(r, columns.subclass)
        if s and s not in subclasses:
            obj, created = Subclass.objects.get_or_create(name=s, clas=classes[c])
            subclasses[s] = obj
            if created:
                logger.info(f"Created Subclass: {obj}")
                total_count += 1

        for treatment_name in split_list(field(r, columns.treatment)):
            if treatment_name not in treatments:
                obj, created = Treatment.objects.get_or_create(name=treatment_name)
                treatments[treatment_name] = obj
                if created:
                    logger.info(f"Created Treatment: {obj}")
                    total_count += 1

        for reference_value in split_list(field(r, columns.reference)):
            if reference_value not in references:
                obj, created = Reference.objects.get_or_create(value=reference_value)
                references[reference_value] = obj
                if created:
                    logger.info(f"Created Reference: {obj}")
                    total_count += 1

        origin = field(r, columns.origin)
        name = field(r, columns.name)

        if field(r, columns.type) == "original":
            originals[origin] = r
        else:
            tps[name] = r

    # Create original compounds
    for origin, r in originals.items():
        obj, created = Compound.objects.get_or_create(
            origin=None,
            clas=classes[field(r, columns.clas)],
            subclass=subclasses[field(r, columns.subclass)],
            type="original",
            mode=True if field(r, columns.mode) == "positive" else False,
            name=field(r, columns.name),
            neutral_formula=r[columns.neutral] or "",
            mz_ion=str(r[columns.mz]) if r[columns.mz] else "",
            smile=r[columns.smile] or "",
            notes=field(r, columns.notes),
        )
        for treatment_name in split_list(field(r, columns.treatment)):
            if treatment_name in treatments and treatments[treatment_name]:
                obj.treatment.add(treatments[treatment_name])
        for reference_value in split_list(field(r, columns.reference)):
            if reference_value in references and references[reference_value]:
                obj.references.add(references[reference_value])

        # Handle FormulaMass models (F1-F10 columns)
        formula_mass_objects = create_formula_mass_models(r, columns)
        for formula_mass_obj in formula_mass_objects:
            obj.formulas.add(formula_mass_obj)
        if formula_mass_objects:
            total_count += len(formula_mass_objects)

        # Generate and save molecule image
        if obj.smile and not skip_images:
            logger.info(
                "Calling generate_and_save_molecule_image for compound: %s",
                obj.name,
            )
            if generate_and_save_molecule_image(obj):
                obj.save()  # Save the compound with the image
                logger.info(f"Generated molecule image for: {obj.name}")

        originals[origin] = obj
        if created:
            logger.info(f"Created Original Compound: {obj}")
            total_count += 1

    # Create TP compounds
    for name, r in tps.items():
        origin_compound = originals.get(field(r, columns.origin))
        obj, created = Compound.objects.get_or_create(
            origin=origin_compound,
            clas=classes[field(r, columns.clas)],
            subclass=subclasses[field(r, columns.subclass)],
            type="TP",
            mode=True if field(r, columns.mode) == "positivo" else False,
            name=name,
            neutral_formula=r[columns.neutral] or "",
            mz_ion=str(r[columns.mz]) if r[columns.mz] else "",
            smile=r[columns.smile] or "",
            notes=field(r, columns.notes),
        )
        for treatment_name in split_list(field(r, columns.treatment)):
            if treatment_name in treatments and treatments[treatment_name]:
                obj.treatment.add(treatments[treatment_name])
        for reference_value in split_list(field(r, columns.reference)):
            if reference_value in references and references[reference_value]:
                obj.references.add(references[reference_value])

        formula_mass_objects = create_formula_mass_models(r, columns)
        for formula_mass_obj in formula_mass_objects:
            obj.formulas.add(formula_mass_obj)
        if formula_mass_objects:
            total_count += len(formula_mass_objects)

        # Generate and save molecule image
        if obj.smile and not skip_images:
            logger.info(
                "Calling generate_and_save_molecule_image for compound: %s",
                obj.name,
            )
            if generate_and_save_molecule_image(obj):
                obj.save()  # Save the compound with the image
                logger.info(f"Generated molecule image for: {obj.name}")

        if created:
            logger.info(f"Created TP Compound: {obj}")
            total_count += 1

    return total_count


def process_pending(max_files=1):
    pending_uploads = ExcelUpload.objects.filter(status="pending").order_by(
        "uploaded_at"
    )[:max_files]

    if not pending_uploads:
        logger.info("No pending Excel uploads found.")
        return

    logger.info(f"Found {len(pending_uploads)} pending upload(s) to process.")

    for upload in pending_uploads:
        logger.info(f"Processing: {upload.file.name}")
        try:
            process_upload(upload)
            logger.info(f"✓ Successfully processed: {upload.file.name}")
        except Exception as e:
            logger.error(
                f"Failed to process upload {upload.id}: {str(e)}", exc_info=True
            )


def process_upload(upload):
    logger.info(f"Starting process_upload for {upload.id}")

    # Check if file exists
    if not upload.file or not os.path.exists(upload.file.path):
        logger.error(f"File not found for upload {upload.id}")
        upload.status = "error"
        upload.error_message = "File not found"
        upload.save()
        raise Exception("File not found")

    logger.info(f"File exists at {upload.file.path}")

    # Update status to processing
    upload.status = "processing"
    upload.save()
    logger.info(f"Status updated to processing")

    try:
        count = import_excel(upload.file.path, upload.clear_existing_data, False)
        upload.records_imported = count
        upload.status = "completed"
        upload.error_message = None  # Clear any previous error
        upload.save()

        # Log import event
        add_user_event(
            upload.uploaded_by,
            "import",
            {
                "filename": upload.file.name,
                "records_imported": count,
                "clear_existing_data": upload.clear_existing_data,
            },
        )

    except Exception as e:
        upload.status = "error"
        upload.error_message = str(e)
        upload.save()
        raise e


def import_excel(path, clear=False, skip_images=False):
    with transaction.atomic():
        try:
            wb = load_workbook(
                path,
                read_only=True,
            )
            ws = wb.active
            if clear:
                clear_data()
            count = import_excel_data(ws, skip_images)
            logger.info(f"Imported {count} records")
            return count

        except FileNotFoundError:
            raise Exception(f'File "{path}" not found')

        except Exception as e:
            raise Exception(f"Error reading Excel file: {e}")
